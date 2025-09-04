from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import io
import os
import traceback

# Reuso do core existente
from src.email_sender import enviar_em_lote
from src.excel_reader import ler_contatos
from src.config import (
    SEND_INTERVAL,
    DAILY_LIMIT as CONF_DAILY_LIMIT,
)

app = Flask(__name__, static_folder="web", static_url_path="/")

# Security headers
@app.after_request
def after_request(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    return response

# CORS configuration (restrito à API) — padrão seguro e compatível com Flask-CORS
CORS(app, resources={r"/api/*": {"origins": "*"}})

# Manipulador específico para requisições OPTIONS para resolver problemas de CORS
@app.route('/api/<path:path>', methods=['OPTIONS'])
def options_handler(path):
    response = app.make_default_options_response()
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, X-Requested-With, Accept'
    return response

@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")

@app.route("/help")
def help_page():
    return send_from_directory(app.static_folder, "help.html")


# Favicon (evita 404/500 em ambientes de produção)
@app.route("/favicon.ico")
def favicon():
    try:
        return send_from_directory(app.static_folder, "favicon.ico")
    except Exception:
        return ("", 204)


# Healthcheck simples para plataformas de deploy
@app.route("/api/health")
def health():
    return jsonify({"status": "ok"}), 200


@app.route("/api/send", methods=["POST", "OPTIONS"])
def api_send():
    # Lidar com requisições OPTIONS para CORS
    if request.method == "OPTIONS":
        response = app.make_default_options_response()
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, X-Requested-With, Accept'
        return response
    try:
        # Arquivo Excel
        if "file" not in request.files:
            return jsonify({"error": "Arquivo 'file' (Excel) é obrigatório"}), 400
        up = request.files["file"]
        data = up.read()
        if not data:
            return jsonify({"error": "Arquivo vazio"}), 400

        # Campos texto
        subject = request.form.get("subject", "").strip()
        text_template = request.form.get("text_template", "").strip()
        html_template = request.form.get("html_template")
        html_template = html_template if html_template and html_template.strip() else None

        # SMTP: obrigatoriamente vindas do formulário (multiusuário).
        smtp_user = (request.form.get("smtp_user") or "").strip()
        smtp_pass = (request.form.get("smtp_pass") or "").strip()
        smtp_server = (request.form.get("smtp_server") or "").strip()
        smtp_port = (request.form.get("smtp_port") or "").strip() or "587"
        from_name = (request.form.get("from_name") or "").strip()
        reply_to = (request.form.get("reply_to") or "").strip()
        from_email = (request.form.get("from_email") or smtp_user).strip()

        # Limite diário
        try:
            daily_limit = int(request.form.get("daily_limit", str(CONF_DAILY_LIMIT)))
        except Exception:
            daily_limit = CONF_DAILY_LIMIT
        daily_limit = max(0, min(100, daily_limit))

        # Em ambientes serverless (Vercel), evite timeouts: cap ajustável e sem espera
        is_prod = os.getenv('VERCEL_ENV') == 'production'
        if is_prod:
            # Permitimos até 5 emails por requisição na Vercel, mas com aviso ao usuário
            vercel_limit = int(os.getenv('VERCEL_EMAIL_LIMIT', '5'))
            daily_limit = min(daily_limit, vercel_limit)

        # Validacoes básicas
        if not subject:
            return jsonify({"error": "'subject' é obrigatório"}), 400
        if not text_template:
            return jsonify({"error": "'text_template' é obrigatório"}), 400
        if not smtp_user or not smtp_pass or not smtp_server or not smtp_port:
            return jsonify({"error": "Credenciais SMTP ausentes. Informe Gmail, Senha de App, servidor e porta."}), 400

        # Ler Excel em memória com openpyxl (sem pandas)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return jsonify({"error": "Não foi possível acessar a primeira planilha"}), 400
            row_iter = ws.iter_rows(values_only=True)
            try:
                first = next(row_iter)
            except StopIteration:
                return jsonify({"error": "Planilha vazia"}), 400
            headers = [str(h).strip() if h is not None else "" for h in first]
            # Normalização de nomes de colunas
            def _norm(s: str) -> str:
                return (
                    str(s)
                    .strip()
                    .lower()
                    .replace("-", "")
                    .replace("_", "")
                    .replace(" ", "")
                )
            colmap = {idx: _norm(name) for idx, name in enumerate(headers)}
            # Descobrir índices de colunas
            idx_email = None
            idx_nome = None
            idx_status = None
            for idx, normed in colmap.items():
                if idx_email is None and "mail" in normed:
                    idx_email = idx
                if normed in {"nome", "name"}:
                    idx_nome = idx
                if normed == "status":
                    idx_status = idx
            if idx_email is None:
                return jsonify({"error": "Planilha inválida: coluna de e-mail ausente. Use 'E-mail' ou 'Email'."}), 400
            # Construir lista de contatos (com Status default)
            contatos_raw = []
            total = 0
            pendentes = 0
            for r in row_iter:
                total += 1
                nome_val = (r[idx_nome] if idx_nome is not None and idx_nome < len(r) else "") if r else ""
                email_val = (r[idx_email] if idx_email is not None and idx_email < len(r) else "") if r else ""
                status_val = (r[idx_status] if idx_status is not None and idx_status < len(r) else "Aguardando") if r else "Aguardando"
                email_str = str(email_val or "").strip()
                status_str = str(status_val or "Aguardando").strip()
                nome_str = str(nome_val or "").strip()
                if status_str != "Contatado" and email_str:
                    pendentes += 1
                    contatos_raw.append({
                        "Nome": nome_str,
                        "E-mail": email_str,
                        "Status": status_str or "Aguardando",
                    })
                    if len(contatos_raw) >= daily_limit:
                        # já coletamos o necessário para envio
                        break
            contatos_envio = contatos_raw[:daily_limit]
            try:
                print(f"[api/send] total={total} pendentes={pendentes} a_enviar={len(contatos_envio)} limite={daily_limit}")
            except Exception:
                pass
        except Exception:
            traceback.print_exc()
            return jsonify({"error": "Falha ao ler o Excel (openpyxl). Verifique o formato (.xlsx) e as colunas."}), 400

        results = []
        # Escolher intervalo adequado ao ambiente
        intervalo_envio = 0 if is_prod else SEND_INTERVAL

        for i, sucesso, email_dest in enviar_em_lote(
            contatos_envio,
            subject,
            text_template,
            intervalo=intervalo_envio,
            anexos=None,
            html_template=html_template,
            smtp_user=smtp_user,
            smtp_pass=smtp_pass,
            smtp_server=smtp_server,
            smtp_port=int(smtp_port),
            from_email=from_email,
            from_name=from_name,
            reply_to=reply_to,
            is_serverless=is_prod,
        ):
            # Captura e-mail retornado pelo gerador (compatível com iteráveis não-DataFrame)
            email = email_dest
            status = "Contatado" if sucesso else "Erro"
            results.append({"index": int(i) if isinstance(i, (int, float)) else str(i), "email": email, "success": bool(sucesso), "status": status})

        summary = {
            "requested": len(contatos_envio),
            "sent_ok": sum(1 for r in results if r["success"]),
            "failed": sum(1 for r in results if not r["success"]),
            "limit": daily_limit,
            "is_vercel": is_prod,
            "vercel_limit": vercel_limit if is_prod else None,
            "message": "Na Vercel, o envio é limitado a {} emails por requisição. Para enviar mais emails, faça múltiplas requisições.".format(vercel_limit) if is_prod else None
        }

        return jsonify({"summary": summary, "results": results})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    # Local development
    app.run(host="0.0.0.0", port=8000, debug=False)
